# -*- coding: utf-8 -*-
"""
Created on December 2024

@author: Ghimciuc Ioan
"""

import os
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from src.reports.gait_angles_report import get_knee_angles, get_hip_angles, get_foot_angles
from src.reports.motion_report import MotionReport
from src.utils.body import Leg
from src.utils.vicon_nexus import Event


def export(report: MotionReport, report_name: str, output_directory: str) -> None:
	workbook = Workbook()

	angles_sheet = workbook.active
	angles_sheet.title = "Gait Angles Report"
	cycle_sheet = workbook.create_sheet("Gait Cycle Report")
	step_sheet = workbook.create_sheet("Gait Step Report")
	all_angles_sheet = workbook.create_sheet("All Gait Angles")

	_add_all_angles_xlsx(all_angles_sheet, report)
	_add_cycle_gait_angles(angles_sheet, report)
	_add_gait_cycles(cycle_sheet, report)
	_add_step_parameters(step_sheet, report)

	if not os.path.exists(output_directory):
		os.makedirs(output_directory)

	output_path = os.path.join(output_directory, f"{report_name}.xlsx")
	workbook.save(output_path)


def _add_cycle_gait_angles(sheet: Worksheet, report: MotionReport):
	sheet.append(["", "Piciorul stâng", "", "", "Piciorul drept", "", "", "Unghiurile etalon"])
	sheet.append(["Frame", "Șold", "Genunchi", "Picior", "Șold", "Genunchi", "Picior", "Șold", "Genunchi", "Picior"])

	for i in range(100):
		left_hip = report.gait_angles_report.left_hip_angles[i]
		left_knee = report.gait_angles_report.left_knee_angles[i]
		left_foot = report.gait_angles_report.left_foot_angles[i]
		right_hip = report.gait_angles_report.right_hip_angles[i]
		right_knee = report.gait_angles_report.right_knee_angles[i]
		right_foot = report.gait_angles_report.right_foot_angles[i]

		ref_hip = report.reference_hip_angles[i]
		ref_knee = report.reference_knee_angles[i]
		ref_foot = report.reference_foot_angles[i]

		sheet.append([i + 1, left_hip, left_knee, left_foot, right_hip, right_knee, right_foot, ref_hip, ref_knee, ref_foot])


def _add_gait_cycles(sheet: Worksheet, report: MotionReport):
	sheet.append(["Faza", "Durata in procente pentru piciorul sâng", "Durata in procente pentru piciorul drept"])
	sheet.append(["Bipodal 1", report.gait_cycle_report.left_phases_percentage_duration["Bipodal 1"], report.gait_cycle_report.right_phases_percentage_duration["Bipodal 1"]])
	sheet.append(["Monopodal", report.gait_cycle_report.left_phases_percentage_duration["Monopodal"], report.gait_cycle_report.right_phases_percentage_duration["Monopodal"]])
	sheet.append(["Bipodal 2", report.gait_cycle_report.left_phases_percentage_duration["Bipodal 2"], report.gait_cycle_report.right_phases_percentage_duration["Bipodal 2"]])
	sheet.append(["Balans", report.gait_cycle_report.left_phases_percentage_duration["Balance"], report.gait_cycle_report.right_phases_percentage_duration["Balance"]])


def _add_step_parameters(sheet: Worksheet, report: MotionReport):
	sheet.append(["Mărimea", "Piciorul stâng", "Picrioul drept", "Unitate"])
	sheet.append(["Viteza", report.gait_step_report.left_step_speed, report.gait_step_report.right_step_speed, "mm/s"])
	sheet.append(["Înălțimea", report.gait_step_report.left_step_height, report.gait_step_report.right_step_height, "mm"])
	sheet.append(["Lungimea", report.gait_step_report.left_step_length, report.gait_step_report.right_step_length, "mm"])
	sheet.append(["Cadența", report.gait_step_report.left_step_cadence, report.gait_step_report.right_step_cadence, "steps/min"])
	sheet.append(["Durata", report.gait_step_report.left_step_frames_duration, report.gait_step_report.right_step_frames_duration, "frames"])
	sheet.append(["Start ciclu", report.left_leg.strike_event.frames[0] + report.start_frame, report.right_leg.strike_event.frames[0] + report.start_frame, "frame"])
	sheet.append(["Stop ciclu", report.left_leg.strike_event.frames[1] + report.start_frame, report.right_leg.strike_event.frames[1] + report.start_frame, "frame"])


def _add_all_angles_xlsx(sheet: Worksheet, report: MotionReport) -> None:
	left_leg_strike_event = Event(report.left_leg.strike_event.context, report.left_leg.strike_event.name, [0, -1], [])
	right_leg_strike_event = Event(report.right_leg.strike_event.context, report.right_leg.strike_event.name, [0, -1], [])

	left_leg = Leg(report.left_leg.side, report.left_leg.markers, left_leg_strike_event, report.left_leg.off_event)
	right_leg = Leg(report.right_leg.side, report.right_leg.markers, right_leg_strike_event, report.right_leg.off_event)

	left_hip_angles = get_hip_angles(left_leg).tolist()
	left_knee_angles = get_knee_angles(left_leg).tolist()
	left_foot_angles = get_foot_angles(left_leg).tolist()

	right_hip_angles = get_hip_angles(right_leg).tolist()
	right_knee_angles = get_knee_angles(right_leg).tolist()
	right_foot_angles = get_foot_angles(right_leg).tolist()

	sheet.append(["", "Piciorul stâng", "", "", "Piciorul drept", "", ""])
	sheet.append(["Frame", "Șold", "Genunchi", "Picior", "Șold", "Genunchi", "Picior"])

	max_len = max(len(left_hip_angles),  len(left_knee_angles), len(left_foot_angles), len(right_hip_angles),  len(right_knee_angles), len(right_foot_angles))

	for i in range(max_len):
		left_hip = left_hip_angles[i] if i < len(left_hip_angles) else ""
		left_knee = left_knee_angles[i] if i < len(left_knee_angles) else ""
		left_foot = left_foot_angles[i] if i < len(left_foot_angles) else ""
		right_hip = right_hip_angles[i] if i < len(right_hip_angles) else ""
		right_knee = right_knee_angles[i] if i < len(right_knee_angles) else ""
		right_foot = right_foot_angles[i] if i < len(right_foot_angles) else ""

		sheet.append([i + report.start_frame, left_hip, left_knee, left_foot, right_hip, right_knee, right_foot])
