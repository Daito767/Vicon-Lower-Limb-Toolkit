import os
from typing import List

from fpdf import FPDF, XPos, YPos
from matplotlib import pyplot as plt
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.reports.gait_angles_report import get_knee_angles, get_hip_angles, get_foot_angles
from src.reports.motion_report import MotionReport
from src.utils.body import Leg
from src.utils.vicon_nexus import Event


def export_plot_leg_angles(angles: dict, reference_angles: dict, phases: List[int], title_prefix: str, output_path: str) -> None:
	"""Generates a plot with knee, foot, and hip angles for a specific leg, overlaying real and reference angles."""
	fig, axs = plt.subplots(3, 1, figsize=(8, 12))

	# Plot hip angles with overlay
	axs[0].plot(angles["hip"], label="Măsurat", color='blue')
	axs[0].plot(reference_angles["hip"], label="Etalon", linestyle="--", color='orange')
	axs[0].set_title(f"Variația amplitudinii unghiului șoldului")
	axs[0].set_xlabel("Ciclul de mers (procente)")
	axs[0].set_ylabel("Unghi (grade)")
	axs[0].legend()
	axs[0].grid(False)

	# Plot knee angles with overlay
	axs[1].plot(angles["knee"], label="Măsurat", color='blue')
	axs[1].plot(reference_angles["knee"], label="Etalon", linestyle="--", color='orange')
	axs[1].set_title(f"Variația amplitudinii unghiului genunchiului")
	axs[1].set_xlabel("Ciclul de mers (procente)")
	axs[1].set_ylabel("Unghi (grade)")
	axs[1].legend()
	axs[1].grid(False)

	# Plot foot angles with overlay
	axs[2].plot(angles["foot"], label="Măsurat", color='blue')
	axs[2].plot(reference_angles["foot"], label="Etalon", linestyle="--", color='orange')
	axs[2].set_title(f"Variația amplitudinii unghiului gleznei")
	axs[2].set_xlabel("Ciclul de mers (procente)")
	axs[2].set_ylabel("Unghi (grade)")
	axs[2].legend()
	axs[2].grid(False)

	rounded_phases = [round(phase) for phase in phases if phase]
	xticks = [x for x in range(0, 101, 5) if all(abs(x - phase) >= 3 for phase in rounded_phases)]

	for ax in axs:
		ax.set_xticks(xticks + rounded_phases)
		for phase in phases:
			ax.axvline(x=phase, color='blue', linestyle='solid', linewidth=0.5)
		for phase in [12, 50, 62]:
			ax.axvline(x=phase, color='orange', linestyle='dashed', linewidth=1, dashes=(5, 10))

	plt.tight_layout()
	plt.savefig(output_path)
	plt.close(fig)


def export_pdf(report: MotionReport, report_name: str, output_directory: str) -> None:
	from src.reports.motion_report import MotionReport
	pdf = FPDF()
	pdf.set_auto_page_break(auto=True, margin=15)
	pdf.set_font("Helvetica", "B", 16)

	left_leg_angles = {
		"knee": report.gait_angles_report.left_knee_angles,
		"foot": report.gait_angles_report.left_foot_angles,
		"hip": report.gait_angles_report.left_hip_angles,
	}
	right_leg_angles = {
		"knee": report.gait_angles_report.right_knee_angles,
		"foot": report.gait_angles_report.right_foot_angles,
		"hip": report.gait_angles_report.right_hip_angles,
	}
	reference_angles = {
		"knee": report.reference_knee_angles,
		"foot": report.reference_foot_angles,
		"hip": report.reference_hip_angles
	}

	left_k = 100 / report.gait_cycle_report.left_total_frame_duration
	right_k = 100 / report.gait_cycle_report.right_total_frame_duration
	left_phases: List[int] = [(phase - report.left_leg.strike_event.frames[0]) * left_k for phase in report.gait_cycle_report.left_cycle_phases.values()]
	right_phases: List[int] = [(phase - report.right_leg.strike_event.frames[0]) * right_k for phase in report.gait_cycle_report.right_cycle_phases.values()]

	# Plot Left Leg Angles
	pdf.add_page()
	left_chart_path = os.path.join(output_directory, "left_leg_angles.png")
	export_plot_leg_angles(left_leg_angles, reference_angles, left_phases, "membrulului inferior stâng", left_chart_path)
	pdf.cell(0, 10, f"Analiza unghiurilor membrului inferior stâng", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
	pdf.image(left_chart_path, x=10, y=25, w=180)

	# Plot Right Leg Angles
	pdf.add_page()
	right_chart_path = os.path.join(output_directory, "right_leg_angles.png")
	export_plot_leg_angles(right_leg_angles, reference_angles, right_phases, "membrulului inferior drept", right_chart_path)
	pdf.cell(0, 10, f"Analiza unghiurilor membrului inferior drept", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align="C")
	pdf.image(right_chart_path, x=10, y=25, w=180)

	# Save the PDF
	pdf_output_path = os.path.join(output_directory, f"{report_name}.pdf")
	pdf.output(pdf_output_path)

	# Clean up images
	os.remove(left_chart_path)
	os.remove(right_chart_path)


def export_xlsx(report: MotionReport, report_name: str, output_directory: str) -> None:
	workbook = Workbook()

	angles_sheet = workbook.active
	angles_sheet.title = "Gait Angles Report"
	cycle_sheet = workbook.create_sheet("Gait Cycle Report")
	step_sheet = workbook.create_sheet("Gait Step Report")
	all_angles_sheet = workbook.create_sheet("All Gait Angles")

	_add_all_angles_xlsx(all_angles_sheet, report)
	_add_cycle_gait_angles(angles_sheet, report)
	_add_gait_cycles(cycle_sheet, report)
	_add_step_parameters(step_sheet, report)

	if not os.path.exists(output_directory):
		os.makedirs(output_directory)

	output_path = os.path.join(output_directory, f"{report_name}.xlsx")
	workbook.save(output_path)


def _add_cycle_gait_angles(sheet: Worksheet, report: MotionReport):
	sheet.append(["", "Piciorul stâng", "", "", "Piciorul drept", "", "", "Unghiurile etalon"])
	sheet.append(["Frame", "Șold", "Genunchi", "Picior", "Șold", "Genunchi", "Picior", "Șold", "Genunchi", "Picior"])

	for i in range(100):
		left_hip = report.gait_angles_report.left_hip_angles[i]
		left_knee = report.gait_angles_report.left_knee_angles[i]
		left_foot = report.gait_angles_report.left_foot_angles[i]
		right_hip = report.gait_angles_report.right_hip_angles[i]
		right_knee = report.gait_angles_report.right_knee_angles[i]
		right_foot = report.gait_angles_report.right_foot_angles[i]

		ref_hip = report.reference_hip_angles[i]
		ref_knee = report.reference_knee_angles[i]
		ref_foot = report.reference_foot_angles[i]

		sheet.append([i + 1, left_hip, left_knee, left_foot, right_hip, right_knee, right_foot, ref_hip, ref_knee, ref_foot])


def _add_gait_cycles(sheet: Worksheet, report: MotionReport):
	sheet.append(["Faza", "Durata in procente pentru piciorul sâng", "Durata in procente pentru piciorul drept"])
	sheet.append(["Bipodal 1", report.gait_cycle_report.left_phases_percentage_duration["Bipodal 1"], report.gait_cycle_report.right_phases_percentage_duration["Bipodal 1"]])
	sheet.append(["Monopodal", report.gait_cycle_report.left_phases_percentage_duration["Monopodal"], report.gait_cycle_report.right_phases_percentage_duration["Monopodal"]])
	sheet.append(["Bipodal 2", report.gait_cycle_report.left_phases_percentage_duration["Bipodal 2"], report.gait_cycle_report.right_phases_percentage_duration["Bipodal 2"]])
	sheet.append(["Balans", report.gait_cycle_report.left_phases_percentage_duration["Balance"], report.gait_cycle_report.right_phases_percentage_duration["Balance"]])


def _add_step_parameters(sheet: Worksheet, report: MotionReport):
	sheet.append(["Mărimea", "Piciorul stâng", "Picrioul drept", "Unitate"])
	sheet.append(["Viteza", report.gait_step_report.left_step_speed, report.gait_step_report.right_step_speed, "mm/s"])
	sheet.append(["Înălțimea", report.gait_step_report.left_step_height, report.gait_step_report.right_step_height, "mm"])
	sheet.append(["Lungimea", report.gait_step_report.left_step_length, report.gait_step_report.right_step_length, "mm"])
	sheet.append(["Cadența", report.gait_step_report.left_step_cadence, report.gait_step_report.right_step_cadence, "steps/min"])
	sheet.append(["Durata", report.gait_step_report.left_step_frames_duration, report.gait_step_report.right_step_frames_duration, "frames"])
	sheet.append(["Start ciclu", report.left_leg.strike_event.frames[0] + report.start_frame, report.right_leg.strike_event.frames[0] + report.start_frame, "frame"])
	sheet.append(["Stop ciclu", report.left_leg.strike_event.frames[1] + report.start_frame, report.right_leg.strike_event.frames[1] + report.start_frame, "frame"])


def _add_all_angles_xlsx(sheet: Worksheet, report: MotionReport) -> None:
	left_leg_strike_event = Event(report.left_leg.strike_event.context, report.left_leg.strike_event.name, [0, -1], [])
	right_leg_strike_event = Event(report.right_leg.strike_event.context, report.right_leg.strike_event.name, [0, -1], [])

	left_leg = Leg(report.left_leg.side, report.left_leg.markers, left_leg_strike_event, report.left_leg.off_event)
	right_leg = Leg(report.right_leg.side, report.right_leg.markers, right_leg_strike_event, report.right_leg.off_event)

	left_hip_angles = get_hip_angles(left_leg).tolist()
	left_knee_angles = get_knee_angles(left_leg).tolist()
	left_foot_angles = get_foot_angles(left_leg).tolist()

	right_hip_angles = get_hip_angles(right_leg).tolist()
	right_knee_angles = get_knee_angles(right_leg).tolist()
	right_foot_angles = get_foot_angles(right_leg).tolist()

	sheet.append(["", "Piciorul stâng", "", "", "Piciorul drept", "", ""])
	sheet.append(["Frame", "Șold", "Genunchi", "Picior", "Șold", "Genunchi", "Picior"])

	max_len = max(len(left_hip_angles),  len(left_knee_angles), len(left_foot_angles), len(right_hip_angles),  len(right_knee_angles), len(right_foot_angles))

	for i in range(max_len):
		left_hip = left_hip_angles[i] if i < len(left_hip_angles) else ""
		left_knee = left_knee_angles[i] if i < len(left_knee_angles) else ""
		left_foot = left_foot_angles[i] if i < len(left_foot_angles) else ""
		right_hip = right_hip_angles[i] if i < len(right_hip_angles) else ""
		right_knee = right_knee_angles[i] if i < len(right_knee_angles) else ""
		right_foot = right_foot_angles[i] if i < len(right_foot_angles) else ""

		sheet.append([i + report.start_frame, left_hip, left_knee, left_foot, right_hip, right_knee, right_foot])
